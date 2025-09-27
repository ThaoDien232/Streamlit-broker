#!/usr/bin/env python3
"""
Check the structure of Prop book.xlsx
"""

import openpyxl
import pandas as pd

try:
    # Load the workbook
    wb = openpyxl.load_workbook('sql/Prop book.xlsx')
    print(f"Worksheets: {wb.sheetnames}")
    
    ws = wb.active
    print(f"Active worksheet: {ws.title}")
    print(f"Max row: {ws.max_row}")
    print(f"Max column: {ws.max_column}")
    
    print("\nHeaders (first row):")
    for i in range(1, min(10, ws.max_column + 1)):
        cell_value = ws.cell(1, i).value
        print(f"  Column {i}: '{cell_value}'")
    
    print("\nSample data (first 3 rows):")
    for r in range(2, min(6, ws.max_row + 1)):
        row_data = []
        for c in range(1, min(8, ws.max_column + 1)):
            row_data.append(ws.cell(r, c).value)
        print(f"  Row {r}: {row_data}")
    
    print("\nLast 3 rows:")
    for r in range(max(2, ws.max_row - 2), ws.max_row + 1):
        row_data = []
        for c in range(1, min(8, ws.max_column + 1)):
            row_data.append(ws.cell(r, c).value)
        print(f"  Row {r}: {row_data}")
    
    # Also try reading with pandas
    print("\n--- Using pandas ---")
    df = pd.read_excel('sql/Prop book.xlsx')
    print(f"Shape: {df.shape}")
    print(f"Columns: {list(df.columns)}")
    print(f"\nFirst 3 rows:")
    print(df.head(3))
    print(f"\nLast 3 rows:")
    print(df.tail(3))
    
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()